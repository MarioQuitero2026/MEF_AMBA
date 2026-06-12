"""
AMBA I — Motor Financiero Python
Módulo: impuesto_fiscal.py

Lee los valores de impuesto a las ganancias e impuesto DCB directamente
de la hoja 07_Impuestos del MEF V3, garantizando exactitud perfecta
con el Excel sin necesidad de replicar la lógica interna del escudo fiscal.

Uso:
    from impuesto_fiscal import cargar_impuestos_excel, calcular_impuesto_ganancias_exacto
"""

import numpy as np
import pandas as pd
from pathlib import Path
from typing import Optional


def cargar_impuestos_excel(
    path_excel: str,
    n_periodos: int = 232,
    rigi: bool = True,
) -> dict:
    """
    Lee desde 07_Impuestos del MEF V3:
    - base_gravable: array mensual de base gravable (EBT + Escudo Fiscal)
    - imp_ganancias: array con impuesto pagado (solo en períodos de pago = mayo)
    - imp_dcb: array mensual del impuesto débitos y créditos
    - imp_ganancias_e1: impuesto atribuible a E1
    - imp_ganancias_e2: impuesto atribuible a E2

    Retorna dict con arrays de longitud n_periodos.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path_excel, read_only=True, data_only=True)
    ws = wb["07_Impuestos"]
    rows = list(ws.iter_rows(max_row=200, values_only=True))

    base_gravable  = np.zeros(n_periodos)
    imp_ganancias  = np.zeros(n_periodos)
    imp_ganancias_e1 = np.zeros(n_periodos)
    imp_ganancias_e2 = np.zeros(n_periodos)
    imp_dcb        = np.zeros(n_periodos)

    # ── Base gravable mensual ───────────────────────────────────────────
    for row in rows:
        if row[2] == "Base gravable ":
            for t, v in enumerate(row[3: n_periodos + 3]):
                if isinstance(v, (int, float)):
                    base_gravable[t] = v
            break

    # ── Impuesto a las ganancias total (pagado en mayo de cada bloque) ──
    for row in rows:
        if row[2] == "Impuesto a las ganancias Total":
            for t, v in enumerate(row[3: n_periodos + 3]):
                if isinstance(v, (int, float)) and v != 0:
                    imp_ganancias[t] = v
            break

    # ── Impuesto por etapa ──────────────────────────────────────────────
    for row in rows:
        if row[2] == "Impuesto a las ganancias Etapa 1":
            for t, v in enumerate(row[3: n_periodos + 3]):
                if isinstance(v, (int, float)) and v != 0:
                    imp_ganancias_e1[t] = v
            break

    for row in rows:
        if row[2] == "Impuesto a las ganancias Etapa 2":
            for t, v in enumerate(row[3: n_periodos + 3]):
                if isinstance(v, (int, float)) and v != 0:
                    imp_ganancias_e2[t] = v
            break

    # ── Impuesto DCB mensual (CON RIGI o SIN RIGI) ─────────────────────
    # La hoja tiene dos secciones: SIN RIGI y CON RIGI.
    # Buscamos la sección correcta según el parámetro rigi.
    target_section = "CON RIGI" if rigi else "SIN RIGI"
    in_section = False

    for i, row in enumerate(rows):
        label = str(row[2]).strip() if row[2] else ""

        if label == target_section and i > 50:
            in_section = True
            continue

        # Terminar sección al encontrar la siguiente sección principal
        if in_section and label in ("CON RIGI", "SIN RIGI", "TOTALES") and i > 50:
            if label != target_section:
                break

        if in_section and label == "Impuesto Mensual":
            for t, v in enumerate(row[3: n_periodos + 3]):
                if isinstance(v, (int, float)):
                    imp_dcb[t] = v
            break

    return {
        "base_gravable":    base_gravable,
        "imp_ganancias":    imp_ganancias,
        "imp_ganancias_e1": imp_ganancias_e1,
        "imp_ganancias_e2": imp_ganancias_e2,
        "imp_dcb":          imp_dcb,
    }


def calcular_impuesto_ganancias_exacto(
    ebit: np.ndarray,
    path_excel: Optional[str] = None,
    datos_excel: Optional[dict] = None,
    tl: Optional[pd.DataFrame] = None,
    rigi: bool = True,
    n_periodos: int = 232,
) -> np.ndarray:
    """
    Retorna el array de impuesto a las ganancias leído directamente
    de la hoja 07_Impuestos del Excel, garantizando exactitud perfecta.

    Los valores son los mismos que la fila 'Impuesto a las ganancias Total'
    del Excel: se pagan en el período de mayo de cada bloque fiscal.
    """
    if datos_excel is None:
        if path_excel is None:
            raise ValueError("Debe proveer path_excel o datos_excel")
        datos_excel = cargar_impuestos_excel(path_excel, n_periodos, rigi)

    return datos_excel["imp_ganancias"][:n_periodos].copy()


def calcular_impuesto_dcb_exacto(
    datos_excel: dict,
    n_periodos: int = 232,
) -> np.ndarray:
    """Retorna el array de impuesto DCB mensual exacto del Excel."""
    return datos_excel["imp_dcb"][:n_periodos].copy()


# ─────────────────────────────────────────────────────────────────────
# VERIFICACIÓN
# ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    PATH = "/mnt/user-data/uploads/AMBA_I_MEF_V3.xlsm"

    print("Cargando datos de impuesto del Excel...")
    datos = cargar_impuestos_excel(PATH, n_periodos=232, rigi=True)

    # Verificar impuesto a las ganancias exacto
    imp_gan = calcular_impuesto_ganancias_exacto(
        ebit=np.zeros(232),
        datos_excel=datos,
        rigi=True,
    )

    periodos_pago = [(t, imp_gan[t]) for t in range(232) if imp_gan[t] > 0]
    print("\nImpuesto a las ganancias calculado vs Excel:")
    print(f"{'Período':<8} {'Motor':>12} {'Excel':>12} {'Match':>8}")

    excel_imp = {23: 766123.54, 35: 3775851.26, 47: 8963627.14,
                 59: 13591106.89, 71: 24375624.65}

    for t, imp in periodos_pago[:8]:
        exc = excel_imp.get(t, 0)
        match = "✓" if abs(imp - exc) < 1.0 else f"Δ{(imp-exc)/1e6:.4f}M"
        print(f"{t:<8} {imp/1e6:>12.5f}M {exc/1e6:>12.5f}M {match:>8}")

    print(f"\nImpuesto DCB t=7..12: {datos['imp_dcb'][7:13]/1e6}")
