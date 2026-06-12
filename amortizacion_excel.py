"""
AMBA I — Motor Financiero Python
Módulo: amortizacion_excel.py
Importa la tabla de amortización directamente del Excel para calibración exacta.
Úsalo cuando necesites replicar exactamente los flujos de deuda del MEF V3.
"""
import numpy as np
import pandas as pd
from datetime import date


def cargar_tabla_amortizacion(path_excel: str) -> pd.DataFrame:
    """
    Lee la hoja '9.a Tabla de Amortización' del MEF V3 y retorna
    un DataFrame con los flujos de deuda período a período.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path_excel, read_only=True, data_only=True)
    ws = wb['9.a Tabla de Amortización']
    rows = list(ws.iter_rows(max_row=300, values_only=True))

    data = []
    for row in rows:
        fecha = row[2]
        if not hasattr(fecha, 'year'):
            continue
        if not isinstance(row[3], (int, float)):
            continue
        data.append({
            'fecha':         pd.Timestamp(fecha),
            'saldo_ini':     row[3] or 0,
            'desembolso':    row[4] or 0,
            'int_generados': row[5] or 0,
            'int_pagados':   row[6] or 0,
            'amortizacion':  row[7] or 0 if len(row) > 7 else 0,
            'saldo_fin':     row[8] or 0 if len(row) > 8 else 0,
        })

    df = pd.DataFrame(data)
    df['servicio_deuda'] = df['amortizacion'] + df['int_pagados']
    return df


def alinear_con_timeline(df_amort: pd.DataFrame,
                          tl: pd.DataFrame) -> pd.DataFrame:
    """
    Alinea la tabla de amortización con el timeline del motor Python,
    mapeando cada fecha al período correspondiente.
    """
    n = len(tl)

    desembolso  = np.zeros(n)
    intereses   = np.zeros(n)
    amort       = np.zeros(n)
    saldo       = np.zeros(n)
    servicio    = np.zeros(n)

    for _, row in df_amort.iterrows():
        # Buscar el período que contiene esta fecha
        fecha_ts = pd.Timestamp(row['fecha'])
        mask = tl['bop'].apply(
            lambda d: d.year == fecha_ts.year and d.month == fecha_ts.month
        )
        if mask.any():
            t = int(tl.loc[mask, 'periodo'].iloc[0])
            desembolso[t]  += row['desembolso']
            intereses[t]   += row['int_generados']
            amort[t]       += row['amortizacion']
            saldo[t]        = row['saldo_fin']
            servicio[t]    += row['servicio_deuda']

    df_out = tl[['periodo', 'bop', 'eop']].copy()
    df_out['desembolso_excel']  = desembolso
    df_out['intereses_excel']   = intereses
    df_out['amortizacion_excel'] = amort
    df_out['saldo_excel']       = saldo
    df_out['servicio_excel']    = servicio

    return df_out


if __name__ == '__main__':
    PATH = '/mnt/user-data/uploads/AMBA_I_MEF_V3.xlsm'
    from timeline import build_timeline
    from params import Params

    p = Params()
    tl = build_timeline(p)
    df_amort = cargar_tabla_amortizacion(PATH)
    df_alineado = alinear_con_timeline(df_amort, tl)

    activos = df_alineado[df_alineado['desembolso_excel'] + df_alineado['amortizacion_excel'] > 0]
    print(activos[['periodo', 'bop', 'desembolso_excel',
                   'intereses_excel', 'amortizacion_excel', 'saldo_excel']].to_string())
    print(f"\nTotal desembolsado: USD {df_alineado['desembolso_excel'].sum()/1e6:.1f}M")
    print(f"Total amortizado:   USD {df_alineado['amortizacion_excel'].sum()/1e6:.1f}M")
    print(f"Saldo final:        USD {df_alineado['saldo_excel'].iloc[-1]:,.0f}")
